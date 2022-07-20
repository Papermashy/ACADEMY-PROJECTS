# NOTE - YOUR gdp.xlsx must be in the same directory as this file in order for it to work.

import openpyxl

# NOTE - You may need to change the name of this file to the name you have given to your own.
wb = openpyxl.load_workbook("my_data_gdp.xlsx")

# get a particular sheet
sh = wb["Data"]
sh2 = wb["Metadata - Countries"]

wb.create_sheet("LA&C")
new_sheet = wb["LA&C"]
new_sheet["A1"] = "Latin America and Caribbean GDP"
new_sheet["A3"] = "Country Code"
new_sheet["B3"] = "Country Name"
new_sheet["C3"] = "1990"
new_sheet["D3"] = "2019"

def sheet2 ():
    country_regions = sh2.iter_rows(min_col=1, max_col=2, min_row=2, values_only=True)
    dict_regions = dict(country_regions)
    return dict_regions

def sheet1 (dict_regions):
    countries = sh.iter_rows(min_col=1, max_col=65, min_row=5, values_only=True)
    for country in countries:
        if (country[1] in dict_regions) and (dict_regions[country[1]] == "Latin America & Caribbean"):
            region_code = dict_regions[country[1]]
            print("{}: {}: {}: {}".format(country[0], region_code, country[33], country[62]))
            new_sheet.append((country[1], country[0], country[34], country[63]))
            wb.save("gdp_modified.xlsx")
            wb.close()

regions_dict = sheet2()
sheet1(regions_dict)